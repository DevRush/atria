#!/usr/bin/env python3
"""Deterministic fixture generator for Atria (hackathon demo data).

Produces, byte-identically on every run:
  fellowship.json       -- cardiology fellowship AY2026-27, per docs/SCHEMA.md
  legacy-schedule.xlsx  -- "last year's" (AY2025-26) messy coordinator workbook,
                           the artifact Claude parses live in the demo

Stdlib + openpyxl only. No wall clock, no unseeded randomness.
Run (see README.md):  .venv/bin/python generate.py

Field names follow docs/SCHEMA.md exactly. Extra top-level keys beyond the
/solve payload (meta, blocks, holidays, assignments) are fixture conveniences;
the web layer strips or consumes them as needed.
"""

from __future__ import annotations

import hashlib
import io
import json
import random
import sys
import zipfile
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
TZ = ZoneInfo("America/New_York")

AY_START = date(2026, 7, 1)
AY_END = date(2027, 6, 30)
N_BLOCKS = 13
BLOCK_DAYS = 28  # block 13 absorbs the year's 365th day (29 days)
SEED = 4711
LEGACY_SEED = 20250701
LEGACY_AY_START = date(2025, 7, 1)

WEEKDAY_CODES = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]

# --------------------------------------------------------------------------
# People — synthetic names only (CLAUDE.md: never real names/schedules).
# id = p_<lastname>. Exactly 3 fellows per weekday clinic day.
# p_okafor (F2, clinicDay THU) is fixed by SCHEMA.md's examples + demo event.
# --------------------------------------------------------------------------
FELLOWS = [
    # (id, name, level, clinicDay)
    ("p_alvarez",    "Sofia Alvarez",    "F1", "MON"),
    ("p_washington", "Jamal Washington", "F1", "TUE"),
    ("p_zhou",       "Wei Zhou",         "F1", "WED"),
    ("p_natarajan",  "Priya Natarajan",  "F1", "THU"),
    ("p_gallagher",  "Erin Gallagher",   "F1", "FRI"),
    ("p_okafor",     "Adaeze Okafor",    "F2", "THU"),
    ("p_cohen",      "Daniel Cohen",     "F2", "MON"),
    ("p_alsayed",    "Fatima Al-Sayed",  "F2", "TUE"),
    ("p_bell",       "Marcus Bell",      "F2", "WED"),
    ("p_suzuki",     "Hana Suzuki",      "F2", "FRI"),
    ("p_ramirez",    "Mateo Ramirez",    "F3", "MON"),
    ("p_park",       "Grace Park",       "F3", "TUE"),
    ("p_adeyemi",    "Tunde Adeyemi",    "F3", "WED"),
    ("p_petrova",    "Anya Petrova",     "F3", "THU"),
    ("p_shah",       "Vikram Shah",      "F3", "FRI"),
]

ROTATIONS = ["CATH", "ECHO", "CCU", "CONSULT", "EP", "NUC", "RESEARCH"]
CORE_ROTATIONS = {"CATH", "ECHO", "CCU", "CONSULT"}
ADVANCED_ROTATIONS = {"EP", "NUC", "RESEARCH"}  # F1s are not eligible

# fellows needed per block, per rotation (RESEARCH absorbs the remainder: 15 - 10)
BLOCK_COVERAGE = {"CATH": 2, "ECHO": 2, "CCU": 2, "CONSULT": 2, "EP": 1, "NUC": 1, "RESEARCH": 5}

SERVICES = [
    # id, name, family, kind, coverage
    ("CATH",     "Cath Lab",           "procedural", "rotation", {"minPerWeekday": 2, "minPerWeekendDay": 0}),
    ("ECHO",     "Echocardiography",   "imaging",    "rotation", {"minPerWeekday": 2, "minPerWeekendDay": 0}),
    ("CCU",      "Coronary Care Unit", "inpatient",  "rotation", {"minPerWeekday": 2, "minPerWeekendDay": 1}),
    ("CONSULT",  "Consult Service",    "consult",    "rotation", {"minPerWeekday": 2, "minPerWeekendDay": 0}),
    ("EP",       "Electrophysiology",  "procedural", "rotation", {"minPerWeekday": 1, "minPerWeekendDay": 0}),
    ("NUC",      "Nuclear Cardiology", "imaging",    "rotation", {"minPerWeekday": 1, "minPerWeekendDay": 0}),
    ("RESEARCH", "Research",           "ambulatory", "rotation", {"minPerWeekday": 0, "minPerWeekendDay": 0}),
    ("CLINIC",   "Continuity Clinic",  "ambulatory", "clinic",   {"minPerWeekday": 3, "minPerWeekendDay": 0}),
    ("CALL",     "Overnight Call",     "inpatient",  "call",     {"minPerWeekday": 1, "minPerWeekendDay": 1}),
    ("JEOP",     "Jeopardy Backup",    "backup",     "jeopardy", {"minPerWeekday": 2, "minPerWeekendDay": 2}),
]

LEGACY_GRADUATED_F3 = [
    ("Rosa Delgado", "F3"), ("Kwame Mensah", "F3"), ("Elena Volkov", "F3"),
    ("Jae-Won Kim", "F3"), ("Omar Haddad", "F3"),
]


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------
def dt(d: date, hour: int, minute: int = 0) -> str:
    """Zone-aware ISO-8601 local datetime (America/New_York, DST-correct)."""
    return datetime(d.year, d.month, d.day, hour, minute, tzinfo=TZ).isoformat()


def daterange(start: date, end: date):
    d = start
    while d <= end:
        yield d
        d += timedelta(days=1)


def nth_weekday(year: int, month: int, weekday: int, n: int) -> date:
    """n-th <weekday> (Mon=0) of a month."""
    d = date(year, month, 1)
    offset = (weekday - d.weekday()) % 7
    return d + timedelta(days=offset + 7 * (n - 1))


def last_weekday(year: int, month: int, weekday: int) -> date:
    if month == 12:
        d = date(year, 12, 31)
    else:
        d = date(year, month + 1, 1) - timedelta(days=1)
    return d - timedelta(days=(d.weekday() - weekday) % 7)


def blocks_for_year(start: date, end: date):
    blocks = []
    for i in range(1, N_BLOCKS + 1):
        b_start = start + timedelta(days=BLOCK_DAYS * (i - 1))
        b_end = start + timedelta(days=BLOCK_DAYS * i - 1) if i < N_BLOCKS else end
        blocks.append({"index": i, "start": b_start.isoformat(), "end": b_end.isoformat()})
    return blocks


def us_holidays() -> list[dict]:
    """US holidays falling inside AY2026-27 (actual dates, hospital's big eight)."""
    hol = [
        (date(2026, 7, 4), "Independence Day"),
        (nth_weekday(2026, 9, 0, 1), "Labor Day"),
        (nth_weekday(2026, 11, 3, 4), "Thanksgiving"),
        (date(2026, 12, 25), "Christmas Day"),
        (date(2027, 1, 1), "New Year's Day"),
        (nth_weekday(2027, 1, 0, 3), "Martin Luther King Jr. Day"),
        (last_weekday(2027, 5, 0), "Memorial Day"),
        (date(2027, 6, 19), "Juneteenth"),
    ]
    return [{"date": d.isoformat(), "name": n} for d, n in sorted(hol)]


# --------------------------------------------------------------------------
# fellowship.json
# --------------------------------------------------------------------------
def build_people() -> list[dict]:
    people = []
    for pid, name, level, clinic_day in FELLOWS:
        if level == "F1":
            rotations = ["CATH", "ECHO", "CCU", "CONSULT"]
        else:  # F2/F3: full rotation eligibility incl. EP, NUC, RESEARCH
            rotations = ["CATH", "ECHO", "CCU", "CONSULT", "EP", "NUC", "RESEARCH"]
        people.append({
            "id": pid, "name": name, "level": level, "fte": 1.0,
            "eligibleServices": rotations + ["CLINIC", "CALL", "JEOP"],
            "clinicDay": clinic_day,
        })
    return people


def build_services() -> list[dict]:
    return [
        {"id": sid, "name": name, "code": sid, "family": family, "kind": kind, "coverage": coverage}
        for sid, name, family, kind, coverage in SERVICES
    ]


def build_slots(blocks: list[dict], holidays: list[dict]) -> list[dict]:
    slots = []
    # 1) block-grain rotation slots: 15 per block (coverage counts above)
    for b in blocks:
        b_start = date.fromisoformat(b["start"])
        b_end = date.fromisoformat(b["end"])
        for svc in ROTATIONS:
            for role in range(1, BLOCK_COVERAGE[svc] + 1):
                slots.append({
                    "id": f"slot_b{b['index']}_{svc.lower()}_{role}",
                    "serviceId": svc,
                    "start": dt(b_start, 7), "end": dt(b_end, 17),
                    "grain": "block", "roleIndex": role,
                })
    # 2) nightly in-house call, 17:00 -> 07:00 (+1d), every night of the year
    for d in daterange(AY_START, AY_END):
        slots.append({
            "id": f"slot_{d.isoformat()}_call_1",
            "serviceId": "CALL",
            "start": dt(d, 17), "end": dt(d + timedelta(days=1), 7),
            "grain": "call-night", "roleIndex": 1,
        })
    # 3) weekly jeopardy, two tiers via roleIndex (1 = activate first, 2 = deep backup)
    #    52 weeks from Jul 1; week 52 extends one extra day to close out the year.
    for w in range(1, 53):
        w_start = AY_START + timedelta(days=7 * (w - 1))
        w_end = w_start + timedelta(days=7) if w < 52 else AY_END + timedelta(days=1)
        for tier in (1, 2):
            slots.append({
                "id": f"slot_w{w:02d}_jeop_{tier}",
                "serviceId": "JEOP",
                "start": dt(w_start, 7), "end": dt(w_end, 7),
                "grain": "week", "roleIndex": tier,
            })
    # 4) continuity-clinic PM half-days: 3 seats per weekday (3 fellows share each
    #    clinicDay), skipped on unit holidays
    holiday_dates = {h["date"] for h in holidays}
    for d in daterange(AY_START, AY_END):
        if d.weekday() >= 5 or d.isoformat() in holiday_dates:
            continue
        for role in (1, 2, 3):
            slots.append({
                "id": f"slot_{d.isoformat()}_clinic_{role}",
                "serviceId": "CLINIC",
                "start": dt(d, 13), "end": dt(d, 17),
                "grain": "halfday", "roleIndex": role,
            })
    return slots


def build_rules() -> list[dict]:
    """~9 catalog rule types; block_requirement expands to one row per (level, service)."""
    rules = []

    def rule(rid, rtype, params, level, tier, scope, source, text, replay_violations=0):
        rules.append({
            "id": rid, "type": rtype, "params": params, "level": level, "tier": tier,
            "scope": scope, "source": source, "text": text, "confirmed": True,
            "replay": {"violationsLastYear": replay_violations},
        })

    # COCATS-ish graduation minimums (block_requirement: N blocks of service S this year)
    rule("r_01", "block_requirement", {"serviceId": "CATH", "minBlocks": 2}, "hard", None, "level:F1",
         "seed:cocats4", "Each first-year fellow completes at least 2 cath lab blocks this year (COCATS Level I trajectory).")
    rule("r_02", "block_requirement", {"serviceId": "ECHO", "minBlocks": 3}, "hard", None, "level:F1",
         "seed:cocats4", "Each first-year fellow completes at least 3 echo blocks this year (COCATS Level I: 3 cumulative months).")
    rule("r_03", "block_requirement", {"serviceId": "CCU", "minBlocks": 2}, "hard", None, "level:F1",
         "seed:program-handbook", "Each first-year fellow completes at least 2 CCU blocks this year.")
    rule("r_04", "block_requirement", {"serviceId": "CATH", "minBlocks": 2}, "hard", None, "level:F2",
         "seed:cocats4", "Each second-year fellow completes at least 2 more cath lab blocks (Level II trajectory).")
    rule("r_05", "block_requirement", {"serviceId": "EP", "minBlocks": 1}, "hard", None, "level:F2",
         "seed:cocats4", "Each second-year fellow completes at least 1 EP block.")
    rule("r_06", "block_requirement", {"serviceId": "RESEARCH", "minBlocks": 4}, "hard", None, "level:F3",
         "seed:program-handbook", "Each third-year fellow gets at least 4 research blocks (research-heavy final year).")

    # coverage + overlay protections
    rule("r_07", "min_coverage", {"serviceId": "CCU", "min": 1, "daily": True}, "hard", None, "all",
         "seed:program-handbook", "CCU must always have at least one fellow on service, every day of the year.")
    rule("r_08", "no_call_before_clinic", {}, "hard", None, "all",
         "excel:25-26 Master!A20", "No overnight call the night before a fellow's continuity clinic day.",
         replay_violations=6)
    rule("r_09", "clinic_day_protected", {"halfDay": "PM"}, "hard", None, "all",
         "seed:cocats4", "Each fellow's fixed weekly continuity-clinic half-day is protected from competing assignments.")

    # ACGME-derived (also hard-coded in the independent validator; blocking = waiver-only override)
    rule("r_10", "min_rest_after_call", {"minHours": 14}, "blocking", None, "all",
         "seed:acgme-cpr-2026", "At least 14 hours free of clinical work after 24h in-house call (ACGME 6.21.a).")
    rule("r_11", "one_in_seven_free", {"averagedOverDays": 28}, "blocking", None, "all",
         "seed:acgme-cpr-2026", "One day in seven free of clinical work, averaged over 4 weeks (ACGME 6.21.b).")

    # equity + spacing (soft tiers)
    rule("r_12", "call_spacing", {"minGapNights": 3}, "soft", "should", "all",
         "seed:program-handbook", "Target q4 call: at least 3 free nights between overnight calls.",
         replay_violations=11)
    rule("r_13", "weekend_equity", {"maxSpread": 2}, "soft", "should", "all",
         "seed:program-handbook", "Weekend call counts stay within 2 of each other across all fellows.",
         replay_violations=3)
    rule("r_14", "holiday_equity", {"maxSpread": 1}, "soft", "must", "all",
         "seed:program-handbook", "Holiday call (unit holiday list) is spread evenly; nobody draws two majors while a peer draws none.",
         replay_violations=1)
    return rules


def build_assignments_and_locks() -> tuple[list[dict], list[dict]]:
    """Two seeded manual assignments so the two locks have referents
    (SCHEMA.md locks point at assignmentId)."""
    assignments = [
        {"id": "a_seed_1", "slotId": "slot_b5_cath_1", "personId": "p_ramirez",
         "status": "draft", "locked": True, "provenance": "manual",
         "createdInVersion": 1, "supersededInVersion": None},
        {"id": "a_seed_2", "slotId": "slot_2026-12-25_call_1", "personId": "p_cohen",
         "status": "draft", "locked": True, "provenance": "manual",
         "createdInVersion": 1, "supersededInVersion": None},
    ]
    locks = [
        {"assignmentId": "a_seed_1", "by": "coordinator",
         "reason": "PD directive: interventional applicant needs cath block before fellowship interviews", "hard": True},
        {"assignmentId": "a_seed_2", "by": "chief_fellow",
         "reason": "Volunteered for Dec 25 call (holiday swap agreement, email 6/30)", "hard": True},
    ]
    return assignments, locks


def build_fixture() -> dict:
    blocks = blocks_for_year(AY_START, AY_END)
    holidays = us_holidays()
    assignments, locks = build_assignments_and_locks()
    return {
        "meta": {
            "name": "Cardiology fellowship demo fixture",
            "academicYear": "2026-27",
            "start": AY_START.isoformat(), "end": AY_END.isoformat(),
            "timezone": "America/New_York",
            "generator": "fixtures/generate.py", "schema": "docs/SCHEMA.md",
        },
        "seed": SEED,
        "blocks": blocks,
        "holidays": holidays,
        "people": build_people(),
        "services": build_services(),
        "slots": build_slots(blocks, holidays),
        "rules": build_rules(),
        "locks": locks,
        "assignments": assignments,
        "absences": [
            {"id": "abs_7", "personId": "p_okafor", "start": "2026-11-13", "end": "2026-11-15",
             "type": "sick", "reasonCode": "OPAQUE-01", "status": "approved"},
        ],
    }


# --------------------------------------------------------------------------
# legacy-schedule.xlsx (AY2025-26) — the messy coordinator workbook
# --------------------------------------------------------------------------
GRID_SHEET = "25-26 Master"
LEGEND_SHEET = "Legend"
CALL_SHEET = "Call Jul 2025"

ROTATION_FILLS = {
    "CATH": "FFC7CE", "ECHO": "BDD7EE", "CCU": "F8CBAD", "CONSULT": "C6E0B4",
    "EP": "D9B3E6", "NUC": "B7E1DC", "RESEARCH": "D9D9D9",
}
CHECK_FILL = "FFFF00"  # "confirm w/ PD" highlight

FOOTNOTE_1 = "* no call w/ clinic Thu"
FOOTNOTE_2 = "† swapped blocks 6/7 with Adeyemi (approved 9/12/25)"


def legacy_roster() -> list[tuple[str, str, str]]:
    """(name, level, clinicDay) for AY2025-26: today's F2s were F1s, today's F3s
    were F2s, plus 5 graduated F3s. Clinic days carried over where known."""
    by_level = {"F1": [], "F2": [], "F3": []}
    for _, name, level, clinic_day in FELLOWS:
        if level == "F2":
            by_level["F1"].append((name, "F1", clinic_day))
        elif level == "F3":
            by_level["F2"].append((name, "F2", clinic_day))
    grad_days = ["MON", "TUE", "WED", "THU", "FRI"]
    for i, (name, _) in enumerate(LEGACY_GRADUATED_F3):
        by_level["F3"].append((name, "F3", grad_days[i]))
    return by_level["F1"] + by_level["F2"] + by_level["F3"]


def build_legacy_grid(roster) -> dict[str, list[str]]:
    """Deterministic 15x13 rotation grid, block coverage exact, F1s core-only."""
    rng = random.Random(LEGACY_SEED)
    template = (["EP", "NUC"] + ["CATH", "CATH", "ECHO", "ECHO", "CCU", "CCU",
                 "CONSULT", "CONSULT"] + ["RESEARCH"] * 5)
    counts = {name: Counter() for name, _, _ in roster}
    last = {name: None for name, _, _ in roster}
    grid = {name: [] for name, _, _ in roster}
    for _ in range(N_BLOCKS):
        remaining = [(name, level) for name, level, _ in roster]
        assigned: dict[str, str] = {}
        for idx, svc in enumerate(template):
            core_left = sum(1 for s in template[idx + 1:] if s in CORE_ROTATIONS)

            def feasible(item):
                name, level = item
                if svc in ADVANCED_ROTATIONS and level == "F1":
                    return False
                f1_left = sum(1 for _, lv in remaining if lv == "F1") - (1 if level == "F1" else 0)
                return f1_left <= core_left  # F1s only fit core slots

            cands = [it for it in remaining if feasible(it)]
            pick = min(cands, key=lambda it: (counts[it[0]][svc],
                                              1 if last[it[0]] == svc else 0,
                                              rng.random()))
            remaining.remove(pick)
            assigned[pick[0]] = svc
        for name, _, _ in roster:
            svc = assigned[name]
            grid[name].append(svc)
            counts[name][svc] += 1
            last[name] = svc
    return grid


def build_legacy_call(roster) -> list[tuple[date, str]]:
    """July 2025 nightly call, seeded round-robin (q15 with 15 fellows)."""
    rng = random.Random(LEGACY_SEED + 1)
    order = [name for name, _, _ in roster]
    rng.shuffle(order)
    return [(date(2025, 7, day), order[(day - 1) % len(order)]) for day in range(1, 32)]


def build_workbook() -> Workbook:
    roster = legacy_roster()
    grid = build_legacy_grid(roster)
    legacy_blocks = blocks_for_year(LEGACY_AY_START, date(2026, 6, 30))

    wb = Workbook()
    fixed = datetime(2026, 7, 1)
    wb.properties.created = fixed
    wb.properties.modified = fixed
    wb.properties.creator = "M. Donnelly (program coordinator)"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ---- grid sheet -------------------------------------------------------
    ws = wb.active
    ws.title = GRID_SHEET
    ws.merge_cells("A1:P1")
    c = ws["A1"]
    c.value = "SMH Cardiology Fellowship — Master Block Schedule AY 2025-2026 (v7 FINAL revised)"
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center")

    seasons = [("Summer", 3, 5), ("Fall", 6, 8), ("Winter", 9, 11), ("Spring", 12, 15)]
    for label, c1, c2 in seasons:
        ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        cell = ws.cell(row=2, column=c1, value=label)
        cell.font = Font(bold=True, italic=True)
        cell.alignment = center

    ws.cell(row=3, column=1, value="Yr").font = Font(bold=True)
    ws.cell(row=3, column=2, value="Fellow").font = Font(bold=True)
    for i, b in enumerate(legacy_blocks):
        s = date.fromisoformat(b["start"])
        e = date.fromisoformat(b["end"])
        cell = ws.cell(row=3, column=3 + i,
                       value=f"B{b['index']}\n{s.month}/{s.day}-{e.month}/{e.day}")
        cell.font = Font(bold=True, size=9)
        cell.alignment = center
    ws.cell(row=3, column=16, value="Notes").font = Font(bold=True)

    # level groups merged in col A; fellows rows 4-18
    for level, r1, r2 in [("F1", 4, 8), ("F2", 9, 13), ("F3", 14, 18)]:
        ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
        cell = ws.cell(row=r1, column=1, value=level)
        cell.font = Font(bold=True)
        cell.alignment = center

    notes = {
        "Adaeze Okafor": "clinic Thu",
        "Anya Petrova": "clinic Thu",
        "Grace Park": "away ACC 3/15-3/17 (blk 10)",
        "Elena Volkov": "grad June — no blk 13 call",
        "Marcus Bell": "check blk 9 w/ PD",
    }
    thu_clinic = {name for name, _, cd in roster if cd == "THU"}
    for r, (name, _, _) in enumerate(roster, start=4):
        display = name + ("*" if name in thu_clinic else "")
        ws.cell(row=r, column=2, value=display).font = Font(size=10)
        for b_i in range(N_BLOCKS):
            svc = grid[name][b_i]
            text = svc
            if name == "Grace Park" and b_i == 5:
                text = svc + "†"          # footnoted swap
            elif name == "Wei Zhou" and b_i == 2:
                text = svc.capitalize()         # mixed case, coordinator-typed
            elif name == "Omar Haddad" and b_i == 7:
                text = svc + " "                # trailing space
            cell = ws.cell(row=r, column=3 + b_i, value=text)
            cell.border = border
            cell.alignment = center
            cell.font = Font(size=9)
            fill = CHECK_FILL if (name == "Marcus Bell" and b_i == 8) else ROTATION_FILLS[svc]
            cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        if name in notes:
            ws.cell(row=r, column=16, value=notes[name]).font = Font(size=9, italic=True)

    ws.cell(row=20, column=1, value=FOOTNOTE_1).font = Font(size=9, italic=True)
    ws.cell(row=21, column=1, value=FOOTNOTE_2).font = Font(size=9, italic=True)

    ws.freeze_panes = "C4"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 20
    for col in range(3, 16):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.column_dimensions["P"].width = 30

    # ---- legend sheet -----------------------------------------------------
    lg = wb.create_sheet(LEGEND_SHEET)
    lg["A1"] = "Color legend (block grid)"
    lg["A1"].font = Font(bold=True)
    legend_names = {
        "CATH": "Cath Lab", "ECHO": "Echo Lab", "CCU": "CCU", "CONSULT": "Consults",
        "EP": "EP", "NUC": "Nuclear", "RESEARCH": "Research / elective",
    }
    for i, (code, label) in enumerate(legend_names.items(), start=3):
        sw = lg.cell(row=i, column=1)
        sw.fill = PatternFill(start_color=ROTATION_FILLS[code], end_color=ROTATION_FILLS[code], fill_type="solid")
        sw.border = border
        lg.cell(row=i, column=2, value=code).font = Font(bold=True)
        lg.cell(row=i, column=3, value=label)
    r = 3 + len(legend_names) + 1
    sw = lg.cell(row=r, column=1)
    sw.fill = PatternFill(start_color=CHECK_FILL, end_color=CHECK_FILL, fill_type="solid")
    sw.border = border
    lg.cell(row=r, column=2, value="YELLOW")
    lg.cell(row=r, column=3, value="confirm w/ PD before publishing")
    lg.column_dimensions["A"].width = 6
    lg.column_dimensions["C"].width = 34

    # ---- call sheet (one month) ------------------------------------------
    cs = wb.create_sheet(CALL_SHEET)
    cs.merge_cells("A1:F1")
    cs["A1"] = "Overnight call — July 2025 (5p-7a)   jeopardy = week at a time"
    cs["A1"].font = Font(bold=True)
    headers = ["Date", "Day", "On Call", "Jeopardy 1st", "Jeopardy 2nd", "Notes"]
    for i, h in enumerate(headers, start=1):
        cell = cs.cell(row=2, column=i, value=h)
        cell.font = Font(bold=True)
        cell.border = border

    call = build_legacy_call(roster)
    rng = random.Random(LEGACY_SEED + 2)
    jeop_pool = [name for name, level, _ in roster if level != "F1"]  # seniors take jeopardy
    call_notes = {
        date(2025, 7, 4): "July 4 — holiday rate",
        date(2025, 7, 18): "swap: Kim ↔ Volkov (email 6/28)",
        date(2025, 7, 23): "*post-call, no clinic",
    }
    for i, (d, name) in enumerate(call):
        r = 3 + i
        cs.cell(row=r, column=1, value=d.strftime("%-m/%-d/%Y"))
        cs.cell(row=r, column=2, value=d.strftime("%a"))
        cs.cell(row=r, column=3, value=name).border = border
        if d in call_notes:
            cs.cell(row=r, column=6, value=call_notes[d]).font = Font(size=9, italic=True)

    # weekly jeopardy pairs, merged vertically across each week's rows (messy but true)
    week_spans, start_day = [], 1
    while start_day <= 31:  # weeks break on Mondays: extend to Sunday or month end
        end_day = start_day
        while end_day < 31 and date(2025, 7, end_day).weekday() != 6:
            end_day += 1
        week_spans.append((start_day, end_day))
        start_day = end_day + 1
    for w_i, (d1, d2) in enumerate(week_spans):
        pair = rng.sample(jeop_pool, 2)
        for col, who in ((4, pair[0]), (5, pair[1])):
            r1, r2 = 3 + d1 - 1, 3 + d2 - 1
            cs.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)
            cell = cs.cell(row=r1, column=col, value=who)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
    for col, w in zip("ABCDEF", (10, 5, 18, 16, 16, 30)):
        cs.column_dimensions[col].width = w
    return wb


def workbook_bytes(wb: Workbook) -> bytes:
    """Serialize + rewrite the zip with fixed entry timestamps -> byte-determinism."""
    raw = io.BytesIO()
    wb.save(raw)
    raw.seek(0)
    out = io.BytesIO()
    with zipfile.ZipFile(raw) as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in sorted(zin.namelist()):
            info = zipfile.ZipInfo(name, date_time=(2026, 7, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o644 << 16
            zout.writestr(info, zin.read(name))
    return out.getvalue()


# --------------------------------------------------------------------------
# consistency checks
# --------------------------------------------------------------------------
class Checker:
    def __init__(self):
        self.passed = 0
        self.failed = 0

    def check(self, ok: bool, label: str):
        if ok:
            self.passed += 1
            print(f"  [ok]   {label}")
        else:
            self.failed += 1
            print(f"  [FAIL] {label}")


def validate_fixture(fx: dict, json_bytes: bytes, ck: Checker):
    print("== fellowship.json checks ==")
    people = fx["people"]
    person_ids = {p["id"] for p in people}
    service_ids = {s["id"] for s in fx["services"]}
    slot_by_id = {s["id"]: s for s in fx["slots"]}
    holiday_dates = {h["date"] for h in fx["holidays"]}

    levels = Counter(p["level"] for p in people)
    ck.check(len(people) == 15 and levels == Counter({"F1": 5, "F2": 5, "F3": 5}),
             f"15 people, 5/5/5 by level ({dict(levels)})")
    ck.check(len(person_ids) == 15 and all(p["id"].startswith("p_") for p in people),
             "person ids unique, p_<lastname> form")
    clinic_days = Counter(p["clinicDay"] for p in people)
    ck.check(all(clinic_days[d] == 3 for d in ["MON", "TUE", "WED", "THU", "FRI"]),
             f"clinicDay: exactly 3 fellows per weekday ({dict(clinic_days)})")
    ck.check(all({"CALL", "JEOP", "CLINIC"} <= set(p["eligibleServices"]) for p in people),
             "every fellow eligible for CALL + JEOP + CLINIC")
    ck.check(all(set(p["eligibleServices"]) <= service_ids for p in people),
             "all eligibleServices reference real services")
    f1_no_adv = all(not (ADVANCED_ROTATIONS & set(p["eligibleServices"]))
                    for p in people if p["level"] == "F1")
    ck.check(f1_no_adv, "rotation eligibility varies by level (F1: no EP/NUC/RESEARCH)")
    okafor = next((p for p in people if p["id"] == "p_okafor"), None)
    ck.check(okafor is not None and okafor["level"] == "F2" and okafor["clinicDay"] == "THU",
             "p_okafor present, F2, clinic THU (schema example preserved)")

    ck.check(service_ids == {"CATH", "ECHO", "CCU", "CONSULT", "EP", "NUC", "RESEARCH",
                             "CLINIC", "CALL", "JEOP"}, "all 10 services present")
    kinds = {s["id"]: s["kind"] for s in fx["services"]}
    ck.check(kinds["CLINIC"] == "clinic" and kinds["CALL"] == "call" and kinds["JEOP"] == "jeopardy"
             and all(kinds[r] == "rotation" for r in ROTATIONS), "service kinds per schema")
    fams = {s["family"] for s in fx["services"]}
    ck.check(fams <= {"procedural", "imaging", "inpatient", "consult", "ambulatory", "backup"},
             "families within the closed 6-family set")

    # blocks
    blocks = fx["blocks"]
    contiguous = all(
        date.fromisoformat(blocks[i + 1]["start"]) == date.fromisoformat(blocks[i]["end"]) + timedelta(days=1)
        for i in range(len(blocks) - 1))
    lengths = [(date.fromisoformat(b["end"]) - date.fromisoformat(b["start"])).days + 1 for b in blocks]
    ck.check(len(blocks) == 13 and blocks[0]["start"] == "2026-07-01" and blocks[-1]["end"] == "2027-06-30"
             and contiguous and lengths[:-1] == [28] * 12 and lengths[-1] == 29,
             "13 contiguous blocks cover Jul 1 2026 - Jun 30 2027 (12x28d + 1x29d = 365d)")

    # slots
    ck.check(len(slot_by_id) == len(fx["slots"]), "slot ids unique")
    ck.check(all(s["serviceId"] in service_ids for s in fx["slots"]),
             "every slot's serviceId exists")
    ck.check(all(s["grain"] in {"block", "week", "day", "call-night", "halfday"} for s in fx["slots"]),
             "every slot grain in schema's closed set")
    parse_ok, tz_ok = True, True
    for s in fx["slots"]:
        try:
            st = datetime.fromisoformat(s["start"])
            en = datetime.fromisoformat(s["end"])
            if st.utcoffset() is None or en.utcoffset() is None or not st < en:
                parse_ok = False
            if st.utcoffset() != datetime(st.year, st.month, st.day, st.hour, tzinfo=TZ).utcoffset():
                tz_ok = False
        except ValueError:
            parse_ok = False
    ck.check(parse_ok, "all slot datetimes ISO-8601, tz-aware, start < end")
    ck.check(tz_ok, "slot offsets match America/New_York incl. DST transitions")

    block_slots = [s for s in fx["slots"] if s["grain"] == "block"]
    per_block = Counter()
    for s in block_slots:
        b_idx = int(s["id"].split("_")[1][1:])
        per_block[(b_idx, s["serviceId"])] += 1
    cov_ok = all(per_block[(b, svc)] == n for b in range(1, 14)
                 for svc, n in BLOCK_COVERAGE.items())
    ck.check(len(block_slots) == 13 * 15 and cov_ok,
             "195 block slots; per-block coverage CATH2/ECHO2/CCU2/CONSULT2/EP1/NUC1/RESEARCH5")
    aligned = all(
        datetime.fromisoformat(s["start"]).date() == date.fromisoformat(blocks[int(s["id"].split("_")[1][1:]) - 1]["start"])
        and datetime.fromisoformat(s["end"]).date() == date.fromisoformat(blocks[int(s["id"].split("_")[1][1:]) - 1]["end"])
        for s in block_slots)
    ck.check(aligned, "block slot start/end dates align with declared block boundaries")

    call_slots = [s for s in fx["slots"] if s["grain"] == "call-night"]
    call_dates = {datetime.fromisoformat(s["start"]).date() for s in call_slots}
    nightly = call_dates == set(daterange(AY_START, AY_END)) and len(call_slots) == 365
    shape_ok = all(datetime.fromisoformat(s["start"]).hour == 17
                   and datetime.fromisoformat(s["end"]).hour == 7
                   and datetime.fromisoformat(s["end"]).date()
                   == datetime.fromisoformat(s["start"]).date() + timedelta(days=1)
                   for s in call_slots)
    ck.check(nightly and shape_ok, "call slots nightly-complete: 365 nights, each 17:00 -> 07:00 (+1d)")

    jeop_slots = [s for s in fx["slots"] if s["serviceId"] == "JEOP"]
    tiers = Counter(s["roleIndex"] for s in jeop_slots)
    j_sorted = sorted((s for s in jeop_slots if s["roleIndex"] == 1),
                      key=lambda s: s["start"])
    j_contig = (datetime.fromisoformat(j_sorted[0]["start"]).date() == AY_START
                and datetime.fromisoformat(j_sorted[-1]["end"]).date() == AY_END + timedelta(days=1)
                and all(j_sorted[i]["end"] == j_sorted[i + 1]["start"] for i in range(len(j_sorted) - 1)))
    ck.check(len(jeop_slots) == 104 and tiers == Counter({1: 52, 2: 52}) and j_contig,
             "jeopardy: 52 weeks x 2 tiers (roleIndex 1/2), contiguous coverage of the whole year")

    clinic_slots = [s for s in fx["slots"] if s["serviceId"] == "CLINIC"]
    exp_days = [d for d in daterange(AY_START, AY_END)
                if d.weekday() < 5 and d.isoformat() not in holiday_dates]
    per_day = Counter(datetime.fromisoformat(s["start"]).date() for s in clinic_slots)
    ck.check(len(clinic_slots) == 3 * len(exp_days)
             and all(per_day[d] == 3 for d in exp_days),
             f"clinic: 3 PM half-day seats on each of {len(exp_days)} non-holiday weekdays")

    # rules
    catalog = {"min_coverage", "max_consecutive_call", "min_rest_after_call", "one_in_seven_free",
               "no_call_before_clinic", "clinic_day_protected", "block_requirement", "max_service_gap",
               "pair_exclusion", "fixed_assignment", "do_not_schedule", "holiday_equity",
               "call_spacing", "weekend_equity", "jeopardy_payback"}
    rules = fx["rules"]
    ck.check(all(r["type"] in catalog for r in rules)
             and len({r["type"] for r in rules}) == 9,
             f"{len(rules)} rules, 9 distinct catalog types, all within the closed v1 set")
    ck.check(all((r["level"] == "soft") == (r["tier"] is not None) for r in rules)
             and all(r["level"] in {"hard", "blocking", "soft"} for r in rules),
             "rule levels valid; tier set iff soft")
    ck.check(all(r["confirmed"] and "violationsLastYear" in r["replay"] for r in rules),
             "all rules confirmed with replay evidence")
    ck.check(all(r["params"].get("serviceId", "CATH") in service_ids for r in rules),
             "rule params reference real services")
    br_demand = Counter()
    for r in rules:
        if r["type"] == "block_requirement":
            br_demand[r["params"]["serviceId"]] += r["params"]["minBlocks"] * 5  # 5 fellows/level
    br_supply_ok = all(br_demand[svc] <= BLOCK_COVERAGE[svc] * 13 for svc in br_demand)
    ck.check(br_supply_ok, "block_requirement demand fits slot supply "
             + str({s: f"{br_demand[s]}/{BLOCK_COVERAGE[s]*13}" for s in sorted(br_demand)}))

    # locks + seeded assignments
    a_by_id = {a["id"]: a for a in fx["assignments"]}
    locks_ok = (len(fx["locks"]) == 2
                and all(lk["assignmentId"] in a_by_id and lk["reason"] and lk["hard"] for lk in fx["locks"]))
    ck.check(locks_ok, "2 locks with reasons, each resolving to a seeded assignment")
    elig = {p["id"]: set(p["eligibleServices"]) for p in people}
    a_ok = all(a["slotId"] in slot_by_id and a["personId"] in person_ids
               and slot_by_id[a["slotId"]]["serviceId"] in elig[a["personId"]]
               and a["locked"] and a["provenance"] == "manual"
               for a in fx["assignments"])
    ck.check(a_ok, "seeded assignments reference real slots/people; person eligible for service")

    # absence (the demo event)
    ab = fx["absences"]
    ck.check(len(ab) == 1 and ab[0] == {
        "id": "abs_7", "personId": "p_okafor", "start": "2026-11-13", "end": "2026-11-15",
        "type": "sick", "reasonCode": "OPAQUE-01", "status": "approved"},
        "exactly one absence: p_okafor sick 2026-11-13 -> 15, approved, opaque reason code")
    ck.check(date.fromisoformat(ab[0]["start"]).weekday() == 4,
             "absence starts on a Friday (spec demo: fellow out Fri-Sun)")

    hols = fx["holidays"]
    ck.check(len(hols) == 8 and all(AY_START <= date.fromisoformat(h["date"]) <= AY_END for h in hols)
             and {h["name"] for h in hols} >= {"Thanksgiving", "Christmas Day", "New Year's Day"},
             "8 US holidays, all inside the academic year")

    size = len(json_bytes)
    ck.check(50_000 < size < 2_000_000,
             f"JSON size sane: {size:,} bytes ({len(fx['slots'])} slots)")


def validate_workbook(xlsx_bytes: bytes, ck: Checker):
    print("== legacy-schedule.xlsx checks ==")
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ck.check(wb.sheetnames == [GRID_SHEET, LEGEND_SHEET, CALL_SHEET],
             f"3 sheets: {wb.sheetnames}")
    ws = wb[GRID_SHEET]
    roster = legacy_roster()
    names = [str(ws.cell(row=r, column=2).value or "").rstrip("*") for r in range(4, 19)]
    ck.check(names == [n for n, _, _ in roster],
             "grid: 15 fellow rows (10 carried over + 5 graduated F3s; current F1s absent)")
    ok_cols = True
    for col in range(3, 16):
        vals = [str(ws.cell(row=r, column=col).value).strip().rstrip("*†").upper()
                for r in range(4, 19)]
        if Counter(vals) != Counter({"RESEARCH": 5, "CATH": 2, "ECHO": 2, "CCU": 2,
                                     "CONSULT": 2, "EP": 1, "NUC": 1}):
            ok_cols = False
    ck.check(ok_cols, "every block column carries exact coverage (2/2/2/2 core + EP + NUC + 5 research)")
    f1_ok = all(str(ws.cell(row=r, column=col).value).strip().rstrip("*†").upper()
                not in ADVANCED_ROTATIONS
                for r in range(4, 9) for col in range(3, 16))
    ck.check(f1_ok, "no then-F1 placed on EP/NUC/RESEARCH (plausible eligibility)")
    merges = {str(m) for m in ws.merged_cells.ranges}
    ck.check("A1:P1" in merges and "A4:A8" in merges and len(merges) >= 8,
             f"grid has merged header/level cells ({len(merges)} merged ranges)")
    fills = {ws.cell(row=r, column=c).fill.start_color.rgb
             for r in range(4, 19) for c in range(3, 16)}
    ck.check(len(fills) >= 8, f"color-coded rotation cells ({len(fills)} distinct fills incl. PD-check yellow)")
    ck.check(ws.cell(row=20, column=1).value == FOOTNOTE_1,
             "footnote '* no call w/ clinic Thu' at 25-26 Master!A20 (r_08 provenance target)")
    starred = [n for r in range(4, 19)
               if str((n := ws.cell(row=r, column=2).value)).endswith("*")]
    ck.check(len(starred) >= 2, f"asterisked fellows present ({len(starred)}: Thu-clinic footnote)")

    lg = wb[LEGEND_SHEET]
    legend_codes = {lg.cell(row=r, column=2).value for r in range(3, 10)}
    ck.check(legend_codes == set(ROTATIONS), "legend sheet maps all 7 rotation colors")

    cs = wb[CALL_SHEET]
    call_names = [cs.cell(row=r, column=3).value for r in range(3, 34)]
    roster_names = {n for n, _, _ in roster}
    ck.check(len(call_names) == 31 and all(n in roster_names for n in call_names),
             "call sheet: 31 July-2025 nights, all names on roster")
    # q-spacing sanity in the legacy call month (round robin => q15)
    gaps_ok = all(call_names[i] != call_names[j]
                  for i in range(31) for j in range(i + 1, min(i + 4, 31)))
    ck.check(gaps_ok, "legacy call month has no fellow within 3 nights of own call (q4+)")
    jeop_merges = [m for m in cs.merged_cells.ranges if m.min_col in (4, 5)]
    ck.check(len(jeop_merges) == 10, "jeopardy columns merged week-at-a-time (5 weeks x 2 tiers)")
    ck.check(len(xlsx_bytes) < 200_000, f"workbook size sane: {len(xlsx_bytes):,} bytes")


# --------------------------------------------------------------------------
# main
# --------------------------------------------------------------------------
def main() -> int:
    fixture = build_fixture()
    json_text = json.dumps(fixture, indent=2, ensure_ascii=False) + "\n"
    json_bytes = json_text.encode("utf-8")
    xlsx_bytes = workbook_bytes(build_workbook())

    ck = Checker()
    validate_fixture(fixture, json_bytes, ck)
    validate_workbook(xlsx_bytes, ck)

    print("== determinism ==")
    json2 = (json.dumps(build_fixture(), indent=2, ensure_ascii=False) + "\n").encode("utf-8")
    xlsx2 = workbook_bytes(build_workbook())
    h = hashlib.sha256
    ck.check(h(json_bytes).hexdigest() == h(json2).hexdigest(),
             f"fellowship.json byte-identical across runs (sha256 {h(json_bytes).hexdigest()[:16]}...)")
    ck.check(h(xlsx_bytes).hexdigest() == h(xlsx2).hexdigest(),
             f"legacy-schedule.xlsx byte-identical across runs (sha256 {h(xlsx_bytes).hexdigest()[:16]}...)")

    if ck.failed:
        print(f"\n{ck.failed} CHECK(S) FAILED ({ck.passed} passed) — nothing written")
        return 1

    (HERE / "fellowship.json").write_bytes(json_bytes)
    (HERE / "legacy-schedule.xlsx").write_bytes(xlsx_bytes)
    print(f"\nALL {ck.passed} CHECKS PASSED")
    print(f"wrote {HERE / 'fellowship.json'} ({len(json_bytes):,} bytes)")
    print(f"wrote {HERE / 'legacy-schedule.xlsx'} ({len(xlsx_bytes):,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
